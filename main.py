from openpyxl import *

# Access sheet from input file
oInputWorkbook = load_workbook("Poorly_Organized_Data_1.xlsx", data_only=True)
InputWorksheet = oInputWorkbook["Grades"]

# Create ouput file, remove default sheet
oOutputWorkbook = Workbook()
oOutputWorkbook.remove(oOutputWorkbook["Sheet"])

# Create a set, iterate through column A, add unique names to set
setClassNames = set()
iRowNumber = 2
location = InputWorksheet[f"A{iRowNumber}"].value
while location:
    setClassNames.add(location)
    iRowNumber += 1
    location = InputWorksheet[f"A{iRowNumber}"].value

# Create sheet
for sClassName in setClassNames:
    oOutputWorkbook.create_sheet(sClassName)

oOutputWorkbook.save(filename="TBD.xlsx")
oOutputWorkbook.close()