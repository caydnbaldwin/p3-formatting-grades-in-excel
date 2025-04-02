import openpyxl
from openpyxl.styles import Font

# Load the original workbook and make easy way to call original ws
wb = openpyxl.load_workbook("Poorly_Organized_Data_1.xlsx")
dataWs = wb["Grades"]

#Change original worksheet so that it has all correct data spilt into correct rows
for row in dataWs.iter_rows(min_row=2, max_row=dataWs.max_row, min_col=1, max_col=dataWs.max_column):
    rowNum = row[0].row
    dataWs[f"E{rowNum}"] = dataWs[f"C{rowNum}"].value
    splitText = dataWs[f"B{rowNum}"].value.split("_")
    dataWs[f"B{rowNum}"] = splitText[0]
    dataWs[f"C{rowNum}"] = splitText[1]
    dataWs[f"D{rowNum}"] = splitText[2]

#Give new updated headers
dataWs["B1"] = "Last Name"
dataWs["C1"] = "First Name"
dataWs["D1"] = "Student ID"
dataWs["E1"] = "Grade"
dataWs["G1"] = "Summary Statistics"
dataWs["H1"] = "Value"

#Make a loop that will cycle through each sheet
for row in dataWs.iter_rows(min_row=2, max_row=dataWs.max_row, min_col=1, max_col=dataWs.max_column):

#Creates new sheets for each of the classes with the class name
        existingSheets = wb.sheetnames
        classNameCheck = row[0].value
        if classNameCheck not in existingSheets:
            new_sheet = wb.copy_worksheet(dataWs)
            new_sheet.title = classNameCheck

#Deletes all data that doesn't pertain to that class
for sheet in wb.worksheets:
    for rowNum2 in range(sheet.max_row, 1, -1):
        if not sheet[f"A{rowNum2}"].value == sheet.title:
            sheet.delete_rows(rowNum2)

#Makes the sheets filterable
    sheet.auto_filter.ref = f"B1:E{sheet.max_row}"

#Finds Summary statistics using Excel functions
    sheet["G2"] = "Highest Grade"
    sheet["H2"] = f"=MAX(E2:E{sheet.max_row})"

    sheet["G3"] = "Lowest Grade"
    sheet["H3"] = f"=MIN(E2:E{sheet.max_row})"

    sheet["G4"] = "Mean Grade"
    sheet["H4"] = f"=AVERAGE(E2:E{sheet.max_row})"

    sheet["G5"] = "Median Grade"
    sheet["H5"] = f"=MEDIAN(E2:E{sheet.max_row})"

    sheet["G6"] = "Number of students in the class"
    sheet["H6"] = f"=COUNT(E2:E{sheet.max_row})"

#Formats the sheet to specifications
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for cell in sheet[1]:
        if cell.value:
            sheet.column_dimensions[cell.column_letter].width = len(cell.value) + 5
    sheet.delete_cols(1)

#deletes original sheet
del wb["Grades"]

#saves new sheet
wb.save("formatted_grades.xlsx")
wb.close()

# print("Hooray! It didn't break!")