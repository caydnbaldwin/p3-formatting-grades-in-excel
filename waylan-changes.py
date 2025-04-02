from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.filters import AutoFilter

# Load input workbook and active sheet
oInputWorkbook = load_workbook("Poorly_Organized_Data_1.xlsx", data_only=True)
ws = oInputWorkbook.active  # Assuming the data is in the first sheet

# Create output workbook and remove default sheet
oOutputWorkbook = Workbook()
oOutputWorkbook.remove(oOutputWorkbook["Sheet"])

# Identify unique class names
setClassNames = set()
for row in ws.iter_rows(min_row=2, values_only=True):
    setClassNames.add(row[0])  # Column A contains class names

# Create worksheets for each class
for sClassName in setClassNames:
    oOutputWorkbook.create_sheet(sClassName)

# Process student data and populate new sheets
for row in ws.iter_rows(min_row=2, values_only=True):
    class_name, student_info, grade = row
    last_name, first_name, student_id = student_info.split("_")

    sheet = oOutputWorkbook[class_name]

    # If sheet is empty, add headers
    if sheet.max_row == 1:
        headers = ["Last Name", "First Name", "Student ID", "Grade"]
        sheet.append(headers)

        # Apply bold formatting to headers
        for col in range(1, 5):
            sheet.cell(row=1, column=col).font = Font(bold=True)

    # Append student data
    sheet.append([last_name, first_name, student_id, grade])

# Apply filters and add summary information
for sheet_name in setClassNames:
    sheet = oOutputWorkbook[sheet_name]
    max_row = sheet.max_row

    # Apply filter to student data
    sheet.auto_filter = AutoFilter(ref=f"A1:D{max_row}")

    # Add summary titles
    summary_titles = ["Highest Grade", "Lowest Grade", "Mean Grade", "Median Grade", "Student Count"]
    for i, title in enumerate(summary_titles):
        sheet[f"F{i+2}"] = title  # Place titles in column F (Rows 2-6)
        sheet[f"F{i+2}"].font = Font(bold=True)  # Bold the summary titles

    # Add formulas in column G
    sheet["G2"] = f"=MAX(D2:D{max_row})"
    sheet["G3"] = f"=MIN(D2:D{max_row})"
    sheet["G4"] = f"=AVERAGE(D2:D{max_row})"
    sheet["G5"] = f"=MEDIAN(D2:D{max_row})"
    sheet["G6"] = f"=COUNTA(D2:D{max_row})"

    # Adjust column widths
    column_headers = ["Last Name", "First Name", "Student ID", "Grade", "Summary", "Values"]
    for col_num, header in enumerate(column_headers, start=1):
        sheet.column_dimensions[get_column_letter(col_num)].width = len(header) + 5

# Save formatted workbook
oOutputWorkbook.save(filename="formatted_grades.xlsx")
oOutputWorkbook.close()
